import os

import openpyxl


class Excel:
    def __init__(self, filename='1.xlsx', data='数据', sheet_index=None, is_create=True):
        dataPath = os.path.abspath(__file__).split('FastApi')[0]
        filename = dataPath + 'FastApi/data/' + filename
        self.excel = openpyxl.Workbook()
        self.filename = filename
        self.data = data
        self.sheet_index = sheet_index
        if not os.path.exists(filename) and is_create:
            self.create_excel()
        self.fn = openpyxl.load_workbook(self.filename)  # 打开文件

    def create_excel(self):
        self.excel.create_sheet(self.data, self.sheet_index)  # 创建表单的方法  可以指定位置
        self.excel.save(self.filename)  # 另存为  创建文件

    def read_date(self):
        list1 = []  # 创建大列表
        sheet = self.fn[self.data]  # 定位到表单
        for i in range(1, sheet.max_row + 1):  # 定位单元格的行
            list = []  # 小列表   每次重新赋值该列表 始终保证每一行为一个列表
            for j in range(1, sheet.max_column + 1):  # 定位单元格的列 
                if sheet.cell(i, j).value:  # 判断是否有None
                    list.append(sheet.cell(i, j).value)  # 插入list列表
            list1.append(list)  # 小列表插入大列表
        return list1  # 返回值

    def write_date(self, row=1, column=1, content=None):
        sheet = self.fn[self.data]
        sheet.cell(row, column, content)  # 写入值2
        self.fn.save(self.filename)
        self.fn.close()


if __name__ == '__main__':
    a = Excel('test.xlsx', 'sheet1', 0)
    a.create_excel()
    a.write_date(1, 1, 'ttttttttt')
    print(a.read_date())
